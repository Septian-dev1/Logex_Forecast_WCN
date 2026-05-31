"""
Export Manager (CSV & Excel)

Modul ini bertanggung jawab untuk mengekspor hasil prediksi 
ke dalam format CSV dan Excel dengan styling profesional.
"""

import numpy as np
import pandas as pd

from .split import compute_insurer_split, build_forecast_detail
from .pivot import build_pivot_dataframe


def export_to_csv(df_raw: pd.DataFrame, df_prophet_master: pd.DataFrame, target_tahun: int, output_path: str = "prophet_forecast_v5.csv"):
    """
    Mengekspor prediksi Prophet berdetail (granular) ke dalam CSV 
    terbatas pada periode komparasi Mei - Desember tahun target.
    """
    print(f"\n📤 Membuat export CSV: {output_path}")
    split_df = compute_insurer_split(df_raw)
    df_forecast_detail = build_forecast_detail(df_raw, df_prophet_master, split_df, None)
    
    if df_forecast_detail.empty:
        print("   ⚠️  Tidak ada data forecast sama sekali untuk diekspor CSV.")
        return
        
    df_out = df_forecast_detail[
        (df_forecast_detail['year'] == target_tahun) &
        (df_forecast_detail['month'] >= 5) &
        (df_forecast_detail['month'] <= 12)
    ].copy()

    if df_out.empty:
        print(f"   ⚠️  Data CSV KOSONG. (Pastikan ada data prophet_master di rentang Mei-Des {target_tahun})")
        return
        
    row_count_before = len(df_out)
    df_out = df_out[df_out['volume'] >= 0.0001].copy()
    row_count_after = len(df_out)
    print(f"   🧹 Cleanup: Membuang {row_count_before - row_count_after:,} baris debu (volume < 0.0001).")
    
    col_map = {
        'product_code': 'product_code', 'specialism_code': 'specialism_code', 'diagnosis_code': 'diagnosis_code',
        'care_type_code': 'care_type', 'payer_code': 'uzovi_code', 'year': 'year', 'month': 'month',
        'volume': 'volume', 'revenue': 'revenue'
    }
    
    for src_col in col_map:
        if src_col not in df_out.columns: df_out[src_col] = np.nan
        
    df_out = df_out[list(col_map.keys())].rename(columns=col_map).copy()
    df_out['volume'] = df_out['volume'].apply(lambda x: f"{float(x):.7f}".replace('.', ',') if pd.notna(x) else '')
    df_out['revenue'] = df_out['revenue'].apply(lambda x: f"{float(x):.6f}".replace('.', ',') if pd.notna(x) else '')
    df_out = df_out[['product_code', 'specialism_code', 'diagnosis_code', 'care_type', 'uzovi_code', 'year', 'month', 'volume', 'revenue']]
    
    df_out.to_csv(output_path, sep=';', index=False, encoding='utf-8-sig')
    print(f"   ✅ Export CSV selesai (Khusus Mei-Des {target_tahun} sebagai Komparasi) → Total baris final: {len(df_out):,}")


def export_to_excel(df_raw: pd.DataFrame, df_prophet_master: pd.DataFrame, specialties_list: list, target_year: int, output_path: str = "prophet_export_v5.xlsx"):
    """
    Menyusun tabel Pivot Dataframe ke dalam file Excel bergaya korporat yang rapih.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    print(f"\n📤 Membuat export Excel HANYA untuk Tab Analisis Pivot: {output_path}")

    all_pivot_frames = []
    for spec in specialties_list:
        pf = build_pivot_dataframe(df_raw, df_prophet_master, target_year, spec)
        if not pf.empty:
            all_pivot_frames.append(pf)
            
    df_analisis = pd.concat(all_pivot_frames, ignore_index=True) if all_pivot_frames else pd.DataFrame()

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if not df_analisis.empty:
            df_export = df_analisis.drop(columns=['Spesialisasi', 'Monitor_Month'], errors='ignore').copy()
            df_export = df_export.rename(columns={'Tipe': 'Row Labels'})

            ws = writer.book.create_sheet('Analisis')
            writer.sheets['Analisis'] = ws
            ws.cell(row=1, column=1, value='year')
            ws.cell(row=1, column=2, value=target_year)
            ws.cell(row=2, column=1, value='Sum of volume Column')

            for col_idx, col_name in enumerate(df_export.columns, start=1):
                ws.cell(row=3, column=col_idx, value=col_name)

            for row_idx, data_row in enumerate(df_export.itertuples(index=False), start=4):
                for col_idx, val in enumerate(data_row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # Styling Blok Excel
            header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
            header_font = Font(color='FFFFFF', bold=True, size=10)
            fill_a      = PatternFill(fill_type='solid', fgColor='DCE6F1')
            fill_b      = PatternFill(fill_type='solid', fgColor='FFFFFF')
            grand_fill  = PatternFill(fill_type='solid', fgColor='BDD7EE')

            ws.cell(row=1, column=1).font = Font(bold=True, size=10)
            ws.cell(row=1, column=2).font = Font(bold=True, size=10)
            ws.cell(row=2, column=1).font = Font(bold=True, size=10)

            n_cols = len(df_export.columns)
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_idx in range(4, ws.max_row + 1):
                label_val  = str(ws.cell(row=row_idx, column=1).value or '')
                is_grand   = 'GRAND TOTAL' in label_val
                is_summary = label_val.startswith('■')
                fill = grand_fill if is_grand else (fill_a if row_idx % 2 == 0 else fill_b)

                for col_idx in range(1, n_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    if is_grand or is_summary: cell.font = Font(bold=True, size=9)
                    else: cell.font = Font(size=9)
                    if col_idx > 1:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right')

            ws.column_dimensions['A'].width = 28
            for col_idx in range(2, n_cols + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 10
                
            ws.freeze_panes = 'A4'
            print(f"   ✅ Tab 'Analisis'      → {len(df_export)} baris, {len(df_export.columns)} kolom")
        else:
            print("   ⚠️  Tidak ada data untuk Analisis Pivot.")

    print(f"\n✅ Export Excel selesai (Tanpa Forecast Data Tab) → {output_path}\n")
